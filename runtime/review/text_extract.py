from __future__ import annotations

import re
import struct
import zipfile
import zlib
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET

from runtime.review.word_markers import contains_wordprocessingml_markers


@dataclass
class TextExtractionResult:
    success: bool
    text: str
    method: str
    error: str | None = None
    raw_markup_text: str | None = None
    meta: dict | None = None
    # PDF 전용 — 텍스트 추출 품질 판정(assess_text_quality 결과). 다른
    # 파일 형식(txt/docx/xlsx/hwp)은 항상 None — 임베디드 텍스트를 그대로
    # 읽으므로 이 품질 문제(스캔/이미지 PDF, CID 폰트 매핑 손상)가 없다.
    quality: "TextQualityAssessment | None" = None


_RX_PAGE_MARKER = re.compile(r"\[페이지\s*(\d+)\]")
_RX_ARTICLE_MARKER = re.compile(r"제\s*\d+\s*조|Article\s+\d+", re.IGNORECASE)
# 2026-09-04 실사례("영업지원 용역계약서") — "제1조/제2조" 대신 "1. 목적",
# "2. 약정 기간"처럼 순수 번호매김 조항 형식을 쓰는 국내 부속합의서/약정서가
# 실제로 흔하다. 이런 문서는 "제N조"/"Article N" 문자열이 전혀 없어도
# 정상적으로 구조화된 계약서이므로, 줄 맨 앞의 "N. 제목" 헤딩이 3개 이상
# 반복되는 경우도 "조항 표지가 있다"는 근거로 인정한다 — 스캔/CID깨짐
# 문서가 우연히 여러 줄에 걸쳐 오름차순 번호로 시작하는 제목 형태를
# 만들어낼 가능성은 매우 낮으므로 오탐 위험 없이 실사례 오탐만 제거한다.
_RX_NUMBERED_CLAUSE_HEADING = re.compile(r"^\s*\d{1,2}\.\s*\S", re.MULTILINE)
_RX_TOKEN = re.compile(r"[가-힣A-Za-z0-9]+")


@dataclass
class TextQualityAssessment:
    total_chars: int
    page_count: int
    per_page_chars: list[int]
    empty_page_count: int
    single_char_token_ratio: float
    has_article_markers: bool
    verdict: str  # "ok" | "low_quality"
    reasons: list[str]

    def to_dict(self) -> dict[str, object]:
        return {
            "total_chars": self.total_chars,
            "page_count": self.page_count,
            "per_page_chars": list(self.per_page_chars),
            "empty_page_count": self.empty_page_count,
            "single_char_token_ratio": round(self.single_char_token_ratio, 3),
            "has_article_markers": self.has_article_markers,
            "verdict": self.verdict,
            "reasons": list(self.reasons),
        }


def assess_text_quality(text: str) -> TextQualityAssessment:
    """PDF에서 추출된 텍스트가 실제로 읽을 수 있는 내용인지 판단한다.

    "글자 수가 0이 아니다"는 실제 계약 내용을 담고 있다는 보장이 되지
    못한다 — 스캔/이미지 PDF나 CID 폰트 매핑이 깨진 PDF는 pdfplumber/
    pymupdf가 "성공"이라며 상당한 길이의 텍스트를 반환하지만, 실제로는
    "연 간 전 략 적 파 트 너 십"처럼 글자마다 공백이 끼거나 무관한 기호가
    섞인 판독 불가능한 문자열인 경우가 있다(2026-09-02 실사례). 이 함수는
    다음 신호를 종합해 그런 경우를 감지한다:
      1. 전체 글자 수가 거의 없음(사실상 빈 PDF)
      2. 페이지 대부분이 비어 있음(이미지 전용 PDF의 전형적 패턴)
      3. 한글/영문/숫자 토큰 중 "한 글자짜리" 비율이 비정상적으로 높음
         (정상적인 한글 문장은 조사가 붙은 여러 글자 단어 위주다 — CID
         매핑이 깨지면 모든 글자 사이에 공백이 들어가 이 비율이 치솟는다)
      4. 상당한 분량인데 "제N조"/"Article N" 조항 표지도, "1. 제목"류
         순수 번호매김 조항 헤딩(3회 이상 반복)도 전혀 없음 — 실제
         계약서/약정서라면 둘 중 하나는 거의 항상 등장하는 구조 신호
    """
    t = text or ""
    total_chars = len(t.strip())
    reasons: list[str] = []

    page_positions = [(int(m.group(1)), m.start()) for m in _RX_PAGE_MARKER.finditer(t)]
    per_page_chars: list[int] = []
    if page_positions:
        bounds = [p[1] for p in page_positions] + [len(t)]
        for i in range(len(page_positions)):
            segment = t[bounds[i]:bounds[i + 1]]
            segment = _RX_PAGE_MARKER.sub("", segment, count=1).strip()
            per_page_chars.append(len(segment))
    page_count = len(per_page_chars)
    empty_page_count = sum(1 for c in per_page_chars if c < 20)

    tokens = _RX_TOKEN.findall(t)
    single_char_tokens = sum(1 for tok in tokens if len(tok) == 1)
    single_char_token_ratio = (single_char_tokens / len(tokens)) if tokens else 0.0

    has_article_markers = bool(_RX_ARTICLE_MARKER.search(t)) or len(_RX_NUMBERED_CLAUSE_HEADING.findall(t)) >= 3

    if total_chars < 50:
        reasons.append("total_chars_near_zero")
    if page_count >= 2 and empty_page_count >= max(1, int(page_count * 0.5)):
        reasons.append("majority_pages_empty")
    if len(tokens) >= 30 and single_char_token_ratio >= 0.35:
        reasons.append("single_char_token_ratio_too_high")
    if total_chars >= 500 and not has_article_markers:
        reasons.append("no_article_markers_despite_length")

    verdict = "low_quality" if reasons else "ok"
    return TextQualityAssessment(
        total_chars=total_chars,
        page_count=page_count,
        per_page_chars=per_page_chars,
        empty_page_count=empty_page_count,
        single_char_token_ratio=single_char_token_ratio,
        has_article_markers=has_article_markers,
        verdict=verdict,
        reasons=reasons,
    )


def _norm_text(s: str) -> str:
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def _contains_wordprocessingml_markers(text: str) -> bool:
    return contains_wordprocessingml_markers(text)


def _strip_zero_width_and_ctrl(text: str) -> str:
    if not text:
        return ""
    s = text.replace("\u200b", "").replace("\u200c", "").replace("\u200d", "").replace("\ufeff", "")
    s = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", s)
    return s



def extract_text_from_file(file_path: Path) -> TextExtractionResult:
    ext = file_path.suffix.lower()
    min_len = 10
    if ext == ".txt":
        try:
            raw = file_path.read_text(encoding="utf-8")
        except Exception:
            raw = file_path.read_text(encoding="utf-8", errors="replace")
        text = _norm_text(_strip_zero_width_and_ctrl(raw))
        if _contains_wordprocessingml_markers(text):
            return TextExtractionResult(False, "", "txt_read", "WordprocessingML markers detected in input text")
        if len(text) < min_len:
            return TextExtractionResult(False, "", "txt_read", "extracted text too short")
        return TextExtractionResult(True, text, "txt_read", None)

    if ext == ".docx":
        try:
            extracted = extract_text_from_docx(file_path)
            text = _norm_text(_strip_zero_width_and_ctrl(extracted["text"]))
            raw_markup_text = extracted.get("raw_markup_text")
            meta = extracted.get("meta")
            if _contains_wordprocessingml_markers(text):
                return TextExtractionResult(
                    False,
                    "",
                    "docx_xml_parser",
                    "WordprocessingML markers detected in extracted text",
                    raw_markup_text=raw_markup_text,
                    meta=meta,
                )
            if len(text) < min_len:
                return TextExtractionResult(
                    False,
                    "",
                    "docx_xml_parser",
                    "extracted text too short",
                    raw_markup_text=raw_markup_text,
                    meta=meta,
                )
            return TextExtractionResult(True, text, "docx_xml_parser", None, raw_markup_text=raw_markup_text, meta=meta)
        except Exception as exc:
            return TextExtractionResult(False, "", "docx_xml_parser", str(exc))

    if ext == ".xlsx":
        try:
            text = extract_text_from_xlsx(file_path)
            text = _norm_text(_strip_zero_width_and_ctrl(text))
            if len(text) < min_len:
                return TextExtractionResult(False, "", "xlsx_reader", "extracted text too short")
            return TextExtractionResult(True, text, "xlsx_reader", None)
        except Exception as exc:
            return TextExtractionResult(False, "", "xlsx_reader", str(exc))

    if ext == ".pdf":
        try:
            extraction = extract_pdf_text_with_quality(file_path)
            text = _norm_text(_strip_zero_width_and_ctrl(extraction.text))
            method_label = f"pdf_reader:{extraction.method}"
            if len(text) < min_len:
                return TextExtractionResult(False, "", method_label, "extracted text too short", quality=extraction.quality)
            if extraction.quality.verdict != "ok":
                # 글자 수는 있지만 스캔/이미지 PDF이거나 CID 폰트 매핑이
                # 깨진 PDF 특유의 판독 불가능한 텍스트다 — native 추출과
                # OCR 모두 실패한 것으로 취급해 성공 처리하지 않는다.
                # 호출부(server.py)가 REVIEW_FAILED_TEXT_EXTRACTION으로
                # 종료해야 한다는 신호로 quality를 그대로 전달한다.
                return TextExtractionResult(
                    False, text, method_label,
                    f"extraction quality low: {', '.join(extraction.quality.reasons)}",
                    quality=extraction.quality,
                )
            return TextExtractionResult(True, text, method_label, None, quality=extraction.quality)
        except Exception as exc:
            return TextExtractionResult(False, "", "pdf_reader", str(exc))

    if ext == ".hwp":
        try:
            text = extract_text_from_hwp(file_path)
            text = _norm_text(_strip_zero_width_and_ctrl(text))
            if len(text) < min_len:
                return TextExtractionResult(False, "", "hwp_reader", "extracted text too short")
            return TextExtractionResult(True, text, "hwp_reader", None)
        except Exception as exc:
            return TextExtractionResult(False, "", "hwp_reader", str(exc))

    return TextExtractionResult(False, "", "unsupported", f"unsupported extension: {ext}")


def extract_text_from_xlsx(file_path: Path) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_lines: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                sheet_lines.append("\t".join(cells))
        if sheet_lines:
            parts.append(f"[시트: {sheet_name}]\n" + "\n".join(sheet_lines))
    wb.close()
    return "\n\n".join(parts)


def extract_text_from_docx(file_path: Path) -> dict[str, object]:
    track_changes_policy = "final"
    parts = ["word/document.xml"]
    with zipfile.ZipFile(file_path, "r") as z:
        names = z.namelist()
        for n in names:
            if re.match(r"word/header\d+\.xml$", n):
                parts.append(n)
            elif re.match(r"word/footer\d+\.xml$", n):
                parts.append(n)
            elif n in ("word/footnotes.xml", "word/endnotes.xml"):
                parts.append(n)
        numbering_xml = z.read("word/numbering.xml") if "word/numbering.xml" in names else None
        numbering_defs = _parse_numbering_xml(numbering_xml) if numbering_xml else None

        texts: list[str] = []
        raw_parts: list[str] = []
        has_track_changes = False
        for part in parts:
            if part not in names:
                continue
            xml_bytes = z.read(part)
            piece = _extract_visible_text_from_word_xml(
                xml_bytes,
                track_changes_policy=track_changes_policy,
                numbering_defs=numbering_defs,
            )
            texts.extend(piece["texts"])
            raw_parts.append(piece["raw_markup_text"])
            has_track_changes = has_track_changes or bool(piece["has_track_changes"])
        return {
            "text": "\n".join([t for t in texts if t]).strip(),
            "raw_markup_text": "\n".join([p for p in raw_parts if p]).strip()[:20000] or None,
            "meta": {
                "has_track_changes": has_track_changes,
                "track_changes_policy": track_changes_policy,
                "parts_included": parts,
            },
        }


def _parse_numbering_xml(xml_bytes: bytes) -> dict[str, object] | None:
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return None
    w_ns = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    ns = {"w": w_ns}
    num_to_abs: dict[str, str] = {}
    abs_levels: dict[str, dict[str, tuple[str, str]]] = {}

    for num in root.findall("w:num", ns):
        num_id = num.get(f"{{{w_ns}}}numId") or num.get("numId")
        abs_el = num.find("w:abstractNumId", ns)
        abs_id = abs_el.get(f"{{{w_ns}}}val") if abs_el is not None else None
        if isinstance(num_id, str) and isinstance(abs_id, str):
            num_to_abs[num_id] = abs_id

    for absn in root.findall("w:abstractNum", ns):
        abs_id = absn.get(f"{{{w_ns}}}abstractNumId") or absn.get("abstractNumId")
        if not isinstance(abs_id, str):
            continue
        levels: dict[str, tuple[str, str]] = {}
        for lvl in absn.findall("w:lvl", ns):
            ilvl = lvl.get(f"{{{w_ns}}}ilvl") or lvl.get("ilvl")
            if not isinstance(ilvl, str):
                continue
            fmt_el = lvl.find("w:numFmt", ns)
            txt_el = lvl.find("w:lvlText", ns)
            num_fmt = (fmt_el.get(f"{{{w_ns}}}val") if fmt_el is not None else None) or "decimal"
            lvl_text = (txt_el.get(f"{{{w_ns}}}val") if txt_el is not None else None) or "%1."
            levels[ilvl] = (str(num_fmt), str(lvl_text))
        abs_levels[abs_id] = levels

    return {"num_to_abs": num_to_abs, "abs_levels": abs_levels}


_HANGUL = ["가", "나", "다", "라", "마", "바", "사", "아", "자", "차", "카", "타", "파", "하"]
_CIRCLED = ["①", "②", "③", "④", "⑤", "⑥", "⑦", "⑧", "⑨", "⑩", "⑪", "⑫", "⑬", "⑭", "⑮", "⑯", "⑰", "⑱", "⑲", "⑳"]


def _to_roman(n: int) -> str:
    vals = [
        (1000, "M"),
        (900, "CM"),
        (500, "D"),
        (400, "CD"),
        (100, "C"),
        (90, "XC"),
        (50, "L"),
        (40, "XL"),
        (10, "X"),
        (9, "IX"),
        (5, "V"),
        (4, "IV"),
        (1, "I"),
    ]
    out = []
    x = n
    for v, sym in vals:
        while x >= v:
            out.append(sym)
            x -= v
    return "".join(out) or str(n)


def _fmt_num(n: int, fmt: str) -> str:
    f = (fmt or "decimal").lower()
    if f in ("decimal", "decimalzero"):
        return str(n)
    if f in ("lowerletter", "loweralpha"):
        return chr(ord("a") + (n - 1) % 26)
    if f in ("upperletter", "upperalpha"):
        return chr(ord("A") + (n - 1) % 26)
    if f == "lowerroman":
        return _to_roman(n).lower()
    if f == "upperroman":
        return _to_roman(n).upper()
    if f in ("koreanhangul", "hangul"):
        return _HANGUL[(n - 1) % len(_HANGUL)]
    if f in ("decimalenclosedcircle", "decimalenclosedcirclechinese"):
        return _CIRCLED[n - 1] if 1 <= n <= len(_CIRCLED) else str(n)
    return str(n)


def _numbering_prefix(
    *,
    numbering_defs: dict[str, object] | None,
    state: dict[str, list[int]],
    num_id: str,
    ilvl: int,
) -> str | None:
    if not num_id:
        return None
    counters = state.get(num_id) or []
    while len(counters) <= ilvl:
        counters.append(0)
    counters[ilvl] += 1
    for j in range(ilvl + 1, len(counters)):
        counters[j] = 0
    for j in range(0, ilvl):
        if counters[j] <= 0:
            counters[j] = 1
    state[num_id] = counters

    abs_id = None
    levels = None
    if isinstance(numbering_defs, dict):
        num_to_abs = numbering_defs.get("num_to_abs")
        abs_levels = numbering_defs.get("abs_levels")
        if isinstance(num_to_abs, dict):
            abs_id = num_to_abs.get(num_id)
        if isinstance(abs_levels, dict) and isinstance(abs_id, str):
            levels = abs_levels.get(abs_id)

    fmt_by_level: dict[int, str] = {}
    lvl_text = None
    if isinstance(levels, dict):
        t = levels.get(str(ilvl))
        if isinstance(t, tuple) and len(t) == 2:
            fmt_by_level[ilvl] = str(t[0])
            lvl_text = str(t[1])
        for k, v in levels.items():
            try:
                lk = int(k)
            except Exception:
                continue
            if isinstance(v, tuple) and len(v) == 2:
                fmt_by_level[lk] = str(v[0])
    if not lvl_text:
        lvl_text = "%1."

    def repl(m: re.Match) -> str:
        idx = int(m.group(1)) - 1
        if idx < 0 or idx >= len(counters):
            return m.group(0)
        fmt = fmt_by_level.get(idx, "decimal")
        return _fmt_num(int(counters[idx] or 0), fmt)

    label = re.sub(r"%(\d)", repl, lvl_text)
    label = label.strip()
    return label if label else None


def _extract_visible_text_from_word_xml(
    xml_bytes: bytes,
    *,
    track_changes_policy: str,
    numbering_defs: dict[str, object] | None,
) -> dict[str, object]:
    raw_markup_text = xml_bytes.decode("utf-8", errors="replace")
    has_track_changes = ("<w:ins" in raw_markup_text) or ("<w:del" in raw_markup_text) or ("<w:delText" in raw_markup_text)
    if track_changes_policy not in ("final", "original"):
        raise ValueError(f"unsupported track_changes_policy: {track_changes_policy}")
    try:
        root = ET.fromstring(xml_bytes)
    except Exception as exc:
        raise ValueError("invalid WordprocessingML XML") from exc

    w_ns = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    ns = {"w": w_ns}
    num_state: dict[str, list[int]] = {}

    def _include_text(*, in_del: bool, in_ins: bool) -> bool:
        if track_changes_policy == "final":
            return not in_del
        return not in_ins

    def walk(node: ET.Element, *, in_del: bool, in_ins: bool, out: list[str]) -> None:
        tag = node.tag
        if tag == f"{{{w_ns}}}del":
            in_del = True
        elif tag == f"{{{w_ns}}}ins":
            in_ins = True

        if tag == f"{{{w_ns}}}t" or tag == f"{{{w_ns}}}delText":
            txt = node.text or ""
            if txt and _include_text(in_del=in_del, in_ins=in_ins):
                out.append(txt)
        elif tag == f"{{{w_ns}}}tab":
            out.append("\t")
        elif tag == f"{{{w_ns}}}br" or tag == f"{{{w_ns}}}cr":
            out.append("\n")
        elif tag == f"{{{w_ns}}}noBreakHyphen":
            out.append("-")

        for ch in list(node):
            walk(ch, in_del=in_del, in_ins=in_ins, out=out)

    paras: list[str] = []
    for p in root.findall(".//w:p", ns):
        buf: list[str] = []
        prefix = None
        ppr = p.find("w:pPr", ns)
        if ppr is not None:
            numpr = ppr.find("w:numPr", ns)
            if numpr is not None:
                ilvl_el = numpr.find("w:ilvl", ns)
                numid_el = numpr.find("w:numId", ns)
                try:
                    ilvl = int(ilvl_el.get(f"{{{w_ns}}}val")) if ilvl_el is not None else None
                except Exception:
                    ilvl = None
                num_id = numid_el.get(f"{{{w_ns}}}val") if numid_el is not None else None
                if isinstance(num_id, str) and ilvl is not None:
                    prefix = _numbering_prefix(numbering_defs=numbering_defs, state=num_state, num_id=num_id, ilvl=ilvl)
        if prefix:
            buf.append(prefix + " ")
        walk(p, in_del=False, in_ins=False, out=buf)
        joined = "".join(buf)
        joined = _strip_zero_width_and_ctrl(joined)
        joined = joined.replace("\t", " ")
        joined = re.sub(r"[ \t]+", " ", joined).strip()
        if joined:
            paras.append(joined)
    return {"texts": paras, "has_track_changes": has_track_changes, "raw_markup_text": raw_markup_text[:20000]}


def _ocr_pdf_pages(file_path: Path) -> str:
    """이미지 기반 PDF를 pymupdf로 렌더링 후 pytesseract OCR."""
    import fitz  # pymupdf
    import pytesseract
    from PIL import Image
    import io

    pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    doc = fitz.open(str(file_path))
    texts: list[str] = []
    for page_num, page in enumerate(doc, 1):
        mat = fitz.Matrix(2.0, 2.0)  # 2x 해상도 (OCR 정확도 향상)
        pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        text = pytesseract.image_to_string(img, lang="kor+eng")
        if text and text.strip():
            texts.append(f"[페이지 {page_num}]\n{text.strip()}")
    doc.close()
    return "\n\n".join(texts)


def _extract_page_text_column_aware(page) -> str:
    """Extract one page's text, splitting it into left/right columns first
    when the page shows a genuine vertical gutter through the middle.

    pdfplumber's plain page.extract_text() orders text primarily by
    vertical position — on a real two-column contract layout, the left and
    right columns sit at the same heights, so their unrelated sentences get
    concatenated onto the same output line (e.g. a "제3조" clause finishing
    in the left column gets glued directly onto "제5조" starting in the
    right column). That silently corrupts article boundaries downstream:
    an article heading buried mid-line is invisible to the line-anchored
    "제N조" heading scanner in clause_extraction.py, so its content gets
    absorbed into whatever article happens to be accumulating at that
    point instead of becoming its own clause.

    Detects a column layout via a wide, page-height-spanning empty band in
    the character x-position histogram (a real gutter, never present on an
    ordinary single-column page where every paragraph spans near the full
    width) and, only then, extracts each column separately and concatenates
    left-column-then-right-column. Falls back to the plain single-stream
    extraction whenever no such gutter is found, so ordinary single-column
    contracts (the overwhelming majority) are completely unaffected.
    """
    plain = page.extract_text() or ""
    chars = page.chars
    if not chars:
        return plain
    width = float(page.width or 0)
    if width <= 0:
        return plain

    bin_count = 100
    bin_w = width / bin_count
    occupied = [False] * bin_count
    for ch in chars:
        try:
            cx = (float(ch["x0"]) + float(ch["x1"])) / 2
        except Exception:
            continue
        b = min(bin_count - 1, max(0, int(cx / bin_w)))
        occupied[b] = True

    # Only look for a gutter within the central band — a real column gap
    # sits roughly in the middle of the page, not near either margin.
    lo, hi = int(bin_count * 0.25), int(bin_count * 0.75)
    best_len, best_start, best_end = 0, -1, -1
    run_start: int | None = None
    for i in range(lo, hi + 1):
        if not occupied[i]:
            if run_start is None:
                run_start = i
        elif run_start is not None:
            if i - run_start > best_len:
                best_len, best_start, best_end = i - run_start, run_start, i
            run_start = None
    if run_start is not None and hi + 1 - run_start > best_len:
        best_len, best_start, best_end = hi + 1 - run_start, run_start, hi + 1

    # Require a meaningfully wide gap — normal word/sentence spacing never
    # leaves a multi-percent-wide band empty across the *entire* page height.
    if best_len * bin_w < width * 0.04:
        return plain

    split_x = (best_start + best_end) / 2.0 * bin_w
    try:
        left = page.within_bbox((0, 0, split_x, page.height))
        right = page.within_bbox((split_x, 0, page.width, page.height))
        left_text = (left.extract_text() or "").strip()
        right_text = (right.extract_text() or "").strip()
    except Exception:
        return plain
    if not left_text or not right_text:
        return plain
    return (left_text + "\n" + right_text).strip()


def extract_text_from_pdf(file_path: Path) -> str:
    """하위 호환용 단순 wrapper — 품질 정보 없이 텍스트만 필요한 기존
    호출부를 위해 남겨둔다. 새 코드는 extract_pdf_text_with_quality()를
    쓸 것."""
    result = extract_pdf_text_with_quality(file_path)
    return result.text


@dataclass
class PdfExtractionResult:
    text: str
    method: str  # "pdfplumber" | "pymupdf" | "ocr_tesseract"
    quality: TextQualityAssessment
    attempts: list[dict[str, object]]


def extract_pdf_text_with_quality(file_path: Path) -> PdfExtractionResult:
    """PDF 텍스트를 추출하되, 글자 수가 0이 아니라는 것만으로 성공 판정을
    내리지 않는다. pdfplumber → pymupdf 순으로 시도하고 각 결과를
    assess_text_quality()로 검사한다 — "ok" 판정이 나오는 첫 결과를 즉시
    채택하되, 상당한 분량(assess 함수가 신뢰할 만큼 페이지가 있음)인데도
    모든 native 방법이 "low_quality"(스캔/이미지 PDF, 또는 CID 폰트 매핑이
    깨진 PDF 특유의 패턴)로 판정되면 OCR(pytesseract, 페이지를 이미지로
    렌더링 후 문자 인식)로 넘어간다. OCR도 low_quality면 OCR 결과를
    그대로 반환하고 verdict는 low_quality로 남겨, 호출부(extract_text_
    from_file)가 REVIEW_FAILED_TEXT_EXTRACTION으로 이어질 신호를 받도록
    한다 — 텍스트가 있어 보인다는 이유로 조용히 통과시키지 않는다."""
    attempts: list[dict[str, object]] = []

    def _try(method: str, fn) -> tuple[str, TextQualityAssessment] | None:
        try:
            text = fn()
        except Exception as exc:
            attempts.append({"method": method, "error": str(exc)})
            return None
        if not text or not text.strip():
            attempts.append({"method": method, "empty": True})
            return None
        q = assess_text_quality(text)
        attempts.append({"method": method, **q.to_dict()})
        return text, q

    def _pdfplumber() -> str:
        import pdfplumber
        texts: list[str] = []
        with pdfplumber.open(str(file_path)) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                t = _extract_page_text_column_aware(page)
                if t and t.strip():
                    texts.append(f"[페이지 {page_num}]\n{t.strip()}")
        return "\n\n".join(texts)

    def _pymupdf() -> str:
        import fitz
        doc = fitz.open(str(file_path))
        try:
            fitz_texts: list[str] = []
            for page_num, page in enumerate(doc, 1):
                t = page.get_text()
                if t and t.strip():
                    fitz_texts.append(f"[페이지 {page_num}]\n{t.strip()}")
            return "\n\n".join(fitz_texts)
        finally:
            doc.close()

    def _ocr() -> str:
        return _ocr_pdf_pages(file_path)

    best: tuple[str, str, TextQualityAssessment] | None = None  # (text, method, quality)
    for method, fn in (("pdfplumber", _pdfplumber), ("pymupdf", _pymupdf)):
        r = _try(method, fn)
        if r is None:
            continue
        text, q = r
        if q.verdict == "ok":
            return PdfExtractionResult(text=text, method=method, quality=q, attempts=attempts)
        if best is None or q.total_chars > best[2].total_chars:
            best = (text, method, q)

    # 두 native 방법 모두 low_quality(또는 아예 텍스트 없음) — 이미지 기반
    # PDF일 가능성이 높다고 보고 OCR로 넘어간다.
    r = _try("ocr_tesseract", _ocr)
    if r is not None:
        text, q = r
        if q.verdict == "ok" or best is None or q.total_chars > best[2].total_chars:
            return PdfExtractionResult(text=text, method="ocr_tesseract", quality=q, attempts=attempts)

    if best is not None:
        text, method, q = best
        return PdfExtractionResult(text=text, method=method, quality=q, attempts=attempts)

    empty_q = assess_text_quality("")
    return PdfExtractionResult(text="", method="none", quality=empty_q, attempts=attempts)


def extract_text_from_hwp(file_path: Path) -> str:
    import olefile

    HWPTAG_PARA_TEXT = 66

    if not olefile.isOleFile(str(file_path)):
        raise ValueError("올바른 HWP 파일이 아닙니다")

    with olefile.OleFileIO(str(file_path)) as ole:
        if not ole.exists("FileHeader"):
            raise ValueError("FileHeader 스트림이 없습니다")

        header_data = ole.openstream("FileHeader").read()
        if not header_data.startswith(b"HWP Document File"):
            raise ValueError("HWP 파일 형식이 아닙니다")

        is_compressed = True
        if len(header_data) >= 40:
            attrs = struct.unpack_from("<I", header_data, 36)[0]
            is_compressed = bool(attrs & 0x01)

        texts: list[str] = []
        section_idx = 0

        while True:
            section_path = f"BodyText/Section{section_idx:04d}"
            if not ole.exists(section_path):
                break

            raw = ole.openstream(section_path).read()
            data = zlib.decompress(raw, -15) if is_compressed else raw

            i = 0
            while i + 4 <= len(data):
                header = struct.unpack_from("<I", data, i)[0]
                tag_id = header & 0x3FF
                size = (header >> 20) & 0xFFF
                i += 4

                if size == 0xFFF:
                    if i + 4 > len(data):
                        break
                    size = struct.unpack_from("<I", data, i)[0]
                    i += 4

                record_end = i + size

                if tag_id == HWPTAG_PARA_TEXT:
                    para_chars: list[str] = []
                    j = i
                    while j + 2 <= record_end:
                        char_code = struct.unpack_from("<H", data, j)[0]
                        j += 2
                        if char_code == 0x000D:
                            break
                        elif char_code == 0x000A:
                            para_chars.append("\n")
                        elif char_code == 0x0009:
                            para_chars.append("\t")
                        elif char_code < 0x0020:
                            j += 16  # inline object: skip remaining bytes
                        else:
                            para_chars.append(chr(char_code))
                    text = "".join(para_chars).strip()
                    if text:
                        texts.append(text)

                i = record_end

            section_idx += 1

        if not texts:
            raise ValueError("텍스트를 추출할 수 없습니다")

        return "\n".join(texts)

